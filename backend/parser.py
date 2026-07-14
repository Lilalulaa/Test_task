"""
===========================================================
parser.py

Универсальный парсер отчетов 1С "Задолженность клиентов по срокам"

Поддерживает:
    • Дистрибьюция
    • Комплектация

Парсер НЕ зависит от фиксированных уровней иерархии.
Структура дерева определяется автоматически по indent.

===========================================================
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional, Dict, List, Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S"
)
logger = logging.getLogger(__name__)


@dataclass(slots=True)
class DocumentRecord:
    """
    Модель одного документа реализации.
    
    Содержит все поля, извлекаемые из отчёта, включая иерархию и финансовые показатели.
    """
    source: str                     # Имя исходного файла
    organization: str               # Организация (корневой уровень)
    sales_department: str           # Отдел продаж (второй уровень)
    department: str                 # Подразделение (отдел: комплектация/дистрибьюция)
    team: Optional[str]             # Команда (при наличии)
    manager: Optional[str]          # Менеджер
    contractor: Optional[str]       # Контрагент
    contract: Optional[str]         # Договор
    document_name: str              # Название документа (тип операции)
    realization_date: Optional[date]    # Дата возникновения (реализации)
    plan_payment_date: Optional[date]   # Дата планового погашения
    days_to_plan: Optional[int]     # Количество дней до планового погашения
    amount: float                   # Сумма реализации
    debt_total: float               # Долг клиента (всего)
    debt_share: float               # Доля долга, %
    overdue: float                  # Просроченная задолженность
    days_overdue: int               # Количество дней просрочки
    our_debt: float                 # Наш долг
    snapshot_date: date             # Дата среза отчёта


@dataclass(slots=True)
class ParserStatistics:
    """Статистика обработки одного Excel-файла."""
    total_rows: int = 0
    skipped_rows: int = 0
    documents: int = 0
    errors: int = 0


# Маппинг колонок Excel: номер колонки (0-based) для каждого поля
DEFAULT_COLUMNS = {
    "name": 0,                  # Название документа
    "realization_date": 4,      # Дата возникновения
    "plan_payment_date": 6,     # Дата планового погашения
    "days_to_plan": 7,          # Дней до планового погашения
    "amount": 8,                # Сумма реализации
    "debt_total": 9,            # Долг клиента (всего)
    "debt_share": 10,           # Доля долга, %
    "overdue": 11,              # Просрочено
    "days_overdue": 12,         # Дней просрочки
    "our_debt": 13              # Наш долг
}


class DebtExcelParser:
    """
    Универсальный парсер отчетов 1С.

    Поддерживает разные структуры дерева.
    Не использует фиксированные уровни — работает только через indent.
    """

    def __init__(self):
        self.columns = DEFAULT_COLUMNS.copy()
        self.stats = ParserStatistics()
        self.snapshot_date: Optional[date] = None
        self.documents: List[DocumentRecord] = []

        # Текущее дерево иерархии: {уровень_отступа: значение}
        # Пример:
        # {
        #     0: "ООО Монолит",
        #     2: "Отдел продаж",
        #     4: "Отдел Дистрибьюции",
        #     6: "Команда",
        #     8: "Менеджер",
        #     10: "Контрагент",
        #     12: "Договор"
        # }
        self.current_path: Dict[int, str] = {}

        # Название текущего исходного файла (используется для определения отдела)
        self.source_name: str = ""


    def parse(self, excel_path: str) -> List[DocumentRecord]:
        """
        Основной метод парсинга одного Excel-файла.

        Args:
            excel_path: Путь к Excel-файлу

        Returns:
            Список извлечённых документов
        """
        self.clear()

        logger.info("=" * 60)
        logger.info("Парсинг файла")
        logger.info(Path(excel_path).name)
        logger.info("=" * 60)

        try:
            workbook = load_workbook(excel_path)
        except Exception as e:
            logger.exception(e)
            self.stats.errors += 1
            return []

        worksheet = workbook.active
        self.source_name = Path(excel_path).stem

        # Определяем дату отчёта
        self.snapshot_date = self._extract_snapshot_date(worksheet)

        # Строка заголовков отчёта (16-я строка, 0-based = 15, но используем 16 для читаемости)
        header_row = 16

        # Парсим строки отчёта
        self._parse_rows(worksheet, header_row)

        workbook.close()

        logger.info("Документов найдено: %s", len(self.documents))
        return self.documents


    # ==========================================================
    # ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ
    # ==========================================================

    def _extract_snapshot_date(self, ws: Worksheet) -> date:
        """
        Извлекает дату отчёта из ячейки с текстом 'Дата отчета:'.

        Поиск выполняется в первых 30 строках.
        Если дата не найдена, используется сегодняшняя дата.

        Returns:
            Дата отчёта
        """
        pattern = re.compile(r"\d{2}\.\d{2}\.\d{4}")

        for row in ws.iter_rows(min_row=1, max_row=30):
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                if "Дата отчета:" in text:
                    match = pattern.search(text)
                    if match:
                        snapshot = datetime.strptime(match.group(), "%d.%m.%Y").date()
                        logger.info("Дата отчета: %s", snapshot)
                        return snapshot

        logger.warning("Дата отчета не найдена. Используется today.")
        return date.today()


    def _get_indent(self, cell: Cell) -> int:
        """
        Определяет уровень отступа строки.

        Приоритет:
        1. Свойство alignment.indent (Excel)
        2. Количество пробелов в начале строки

        Returns:
            Уровень отступа, -1 если значение пустое
        """
        if cell.value is None:
            return -1

        alignment = cell.alignment
        if alignment is not None and alignment.indent is not None:
            return int(alignment.indent)

        text = str(cell.value)
        spaces = len(text) - len(text.lstrip())
        return spaces


    def _to_float(self, value: Any) -> float:
        """Безопасное преобразование в float."""
        if value is None:
            return 0.0
        try:
            return float(value)
        except Exception:
            return 0.0


    def _to_int(self, value: Any) -> int:
        """Безопасное преобразование в int."""
        if value is None:
            return 0
        try:
            return int(value)
        except Exception:
            return 0


    def _to_date(self, value: Any) -> Optional[date]:
        """
        Безопасное преобразование в date.

        Поддерживает: datetime, date, строку в формате DD.MM.YYYY
        """
        if value is None:
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        try:
            return datetime.strptime(str(value), "%d.%m.%Y").date()
        except Exception:
            return None


    # ==========================================================
    # ОСНОВНОЙ ЦИКЛ ПАРСИНГА
    # ==========================================================

    def _parse_rows(self, ws: Worksheet, header_row: int) -> None:
        """
        Основной цикл обработки строк Excel.

        Args:
            ws: Рабочий лист Excel
            header_row: Номер строки заголовков (1-based)
        """
        total_processed = 0
        total_saved = 0

        # Начинаем со строки после заголовков
        for row in ws.iter_rows(min_row=header_row + 1):
            self.stats.total_rows += 1

            first_cell = row[self.columns["name"]]
            if first_cell.value is None:
                continue

            name = str(first_cell.value).strip()
            indent = self._get_indent(first_cell)

            # Обновляем дерево иерархии
            self._update_tree(indent, name)

            # Проверяем, является ли строка документом
            if not self._is_document_row(row):
                continue

            total_processed += 1

            # Строим документ
            document = self._build_document(row)

            if document and self._validate_document(document):
                self.documents.append(document)
                self.stats.documents += 1
                total_saved += 1
            else:
                self.stats.skipped_rows += 1

        # Диагностика
        logger.info(
            f"Файл {self.source_name}: обработано {total_processed} документов, "
            f"сохранено {total_saved}"
        )


    def _update_tree(self, indent: int, value: str) -> None:
        """
        Обновляет текущее дерево иерархии.

        Принцип:
        - Если уровень отступа уменьшился — удаляем все дочерние уровни
        - Добавляем или обновляем текущий уровень

        Args:
            indent: Уровень отступа текущей строки
            value: Значение текущей строки
        """
        # Удаляем все уровни глубже текущего
        for level in sorted(list(self.current_path.keys()), reverse=True):
            if level >= indent:
                del self.current_path[level]

        # Сохраняем текущий уровень
        self.current_path[indent] = value


    def _is_document_row(self, row) -> bool:
        """
        Проверяет, является ли строка документом реализации.

        Критерии:
        1. Название содержит ключевые паттерны (Реализация, Поступление, Корректировка и т.д.)
        2. Имеет хотя бы один ненулевой финансовый показатель
        3. Имеет дату возникновения или планового погашения (для не-корректировок)

        Returns:
            True если строка является документом
        """
        first_cell = row[self.columns["name"]]
        if first_cell.value is None:
            return False

        name = str(first_cell.value).strip()

        # Список паттернов названий документов
        document_patterns = [
            "Реализация товаров и услуг",
            "Поступление безналичных ДС",
            "Корректировка реализации",
            "Корректировка задолженности",
            "Приходный кассовый ордер",
            "Возврат товаров от клиента",
            "Взаимозачет задолженности",
            "Эквайринговая операция",
        ]

        is_document_by_name = any(pattern in name for pattern in document_patterns)

        if is_document_by_name:
            amount = self._to_float(row[self.columns["amount"]].value)
            debt_total = self._to_float(row[self.columns["debt_total"]].value)
            overdue = self._to_float(row[self.columns["overdue"]].value)
            our_debt = self._to_float(row[self.columns["our_debt"]].value)

            return (
                amount != 0 or
                debt_total != 0 or
                overdue != 0 or
                our_debt != 0
            )

        # Альтернативная проверка: наличие дат и финансовых показателей
        has_realization_date = row[self.columns["realization_date"]].value is not None
        has_plan_date = row[self.columns["plan_payment_date"]].value is not None

        if has_realization_date or has_plan_date:
            # Исключаем договоры с датами
            if "Договор" in name and "Реализация" not in name:
                return False

            amount = self._to_float(row[self.columns["amount"]].value)
            debt_total = self._to_float(row[self.columns["debt_total"]].value)
            overdue = self._to_float(row[self.columns["overdue"]].value)

            return (amount != 0 or debt_total != 0 or overdue != 0)

        return False


    def _get_tree_path(self) -> list[str]:
        """
        Возвращает путь от корня до текущего документа.

        Returns:
            Список значений от корня до текущего уровня
            Пример: ["ООО", "Отдел продаж", "Отдел", "Команда", "Менеджер", "Контрагент", "Договор"]
        """
        return [self.current_path[level] for level in sorted(self.current_path.keys())]


    def _extract_hierarchy(self) -> Dict[str, Optional[str]]:
        """
        Извлекает иерархию из текущего пути.

        Особенности:
        - Отдел определяется по имени файла (source_name)
        - Остальные уровни извлекаются из пути

        Returns:
            Словарь с полями: organization, sales_department, department, team, manager, contractor, contract
        """
        path = self._get_tree_path()

        result = {
            "organization": None,
            "sales_department": None,
            "department": None,
            "team": None,
            "manager": None,
            "contractor": None,
            "contract": None
        }

        if len(path) < 3:
            return result

        # Организация (уровень 0)
        result["organization"] = path[0]
        result["sales_department"] = path[1] if len(path) > 1 else None

        # Отдел (определяется по имени файла)
        if "complect" in self.source_name.lower():
            result["department"] = "Отдел комплектации"
        elif "distribution" in self.source_name.lower():
            result["department"] = "Отдел Дистрибьюции"
        else:
            # Fallback: поиск по ключевым словам в пути
            for item in path:
                if "Дистрибьюция" in item:
                    result["department"] = "Отдел Дистрибьюции"
                    break
                elif "комплектация" in item.lower():
                    result["department"] = "Отдел комплектации"
                    break

            # Если всё ещё None — берём третий элемент пути
            if result["department"] is None and len(path) > 2:
                result["department"] = path[2]

        # Определяем индекс отдела в пути
        dept_index = None
        for i, item in enumerate(path):
            if item == result["department"]:
                dept_index = i
                break

        # Берём остаток пути после отдела
        if dept_index is not None:
            tail = path[dept_index + 1:]
        else:
            tail = path[3:] if len(path) > 3 else []

        if tail:
            # Команда (первый элемент после отдела)
            result["team"] = tail[0] if tail else None

            # Менеджер (ищем по ключевым словам)
            manager_found = False
            for item in tail:
                if "менеджер" in item.lower() or "Калинина" in item or "Романов" in item or "Кусакин" in item:
                    result["manager"] = item
                    manager_found = True
                    break

            # Если менеджер не найден — пытаемся определить по позиции
            if not manager_found and len(tail) >= 2:
                for i, item in enumerate(tail):
                    if "ООО" not in item and "ИП" not in item and "АО" not in item and "Договор" not in item:
                        if i < len(tail) - 1:
                            result["manager"] = item
                            break

            # Контрагент (ищем по ключевым словам)
            for item in tail:
                if "ООО" in item or "ИП" in item or "АО" in item or "физлицо" in item.lower():
                    if item != result["manager"] and item != result["team"]:
                        result["contractor"] = item
                        break

            # Если контрагент не найден — берём предпоследний элемент
            if result["contractor"] is None and len(tail) >= 2:
                for i, item in enumerate(tail):
                    if "Договор" not in item and "ЗМР" not in item:
                        if i < len(tail) - 1:
                            result["contractor"] = item
                            break

            # Договор (последний элемент)
            if tail:
                last = tail[-1]
                if "Договор" in last or "ЗМР" in last:
                    result["contract"] = last
                elif len(tail) > 1:
                    result["contract"] = tail[-2]

        return result


    def _build_document(self, row) -> Optional[DocumentRecord]:
        """
        Создаёт объект DocumentRecord из текущей строки Excel.

        Returns:
            DocumentRecord или None, если не удалось создать
        """
        # Получаем текущую иерархию
        hierarchy = self._extract_hierarchy()

        # Проверка обязательных полей
        if hierarchy["organization"] is None:
            return None

        if hierarchy["department"] is None:
            return None

        # Название документа
        document_name = str(row[self.columns["name"]].value).strip()

        # Читаем финансовые показатели
        realization_date = self._to_date(row[self.columns["realization_date"]].value)
        plan_payment_date = self._to_date(row[self.columns["plan_payment_date"]].value)
        days_to_plan = self._to_int(row[self.columns["days_to_plan"]].value)

        amount = self._to_float(row[self.columns["amount"]].value)
        debt_total = self._to_float(row[self.columns["debt_total"]].value)
        debt_share = self._to_float(row[self.columns["debt_share"]].value)
        overdue = self._to_float(row[self.columns["overdue"]].value)
        days_overdue = self._to_int(row[self.columns["days_overdue"]].value)
        our_debt = self._to_float(row[self.columns["our_debt"]].value)

        # Создаём объект документа
        document = DocumentRecord(
            source=self.source_name,
            organization=hierarchy["organization"],
            sales_department=hierarchy["sales_department"],
            department=hierarchy["department"],
            team=hierarchy["team"],
            manager=hierarchy["manager"],
            contractor=hierarchy["contractor"],
            contract=hierarchy["contract"],
            document_name=document_name,
            realization_date=realization_date,
            plan_payment_date=plan_payment_date,
            days_to_plan=days_to_plan,
            amount=amount,
            debt_total=debt_total,
            debt_share=debt_share,
            overdue=overdue,
            days_overdue=days_overdue,
            our_debt=our_debt,
            snapshot_date=self.snapshot_date
        )

        if not self._validate_document(document):
            self.stats.skipped_rows += 1
            return None

        return document


    def _validate_document(self, document: DocumentRecord) -> bool:
        """
        Проверяет корректность документа перед сохранением.

        Критерии:
        - Обязательные поля заполнены (organization, department, document_name)
        - Есть хотя бы один ненулевой финансовый показатель
        - Есть дата (или это корректировка)

        Returns:
            True если документ валиден
        """
        # Обязательные поля
        if not document.organization:
            return False
        if not document.department:
            return False
        if not document.document_name:
            return False

        # Наличие финансовых показателей
        has_value = (
            document.amount != 0 or
            document.debt_total != 0 or
            document.overdue != 0 or
            document.our_debt != 0
        )
        if not has_value:
            return False

        # Наличие даты (для корректировок допускается отсутствие)
        has_date = (
            document.realization_date is not None or
            document.plan_payment_date is not None
        )
        if not has_date and "Корректировка" not in document.document_name:
            return False

        return True


    def clear(self) -> None:
        """
        Полностью очищает состояние парсера.

        Используется перед обработкой следующего Excel-файла.
        """
        self.documents.clear()
        self.current_path.clear()
        self.stats = ParserStatistics()
        self.snapshot_date = None
        self.source_name = ""


    def print_statistics(self) -> None:
        """Выводит статистику после завершения парсинга."""
        logger.info("=" * 60)
        logger.info("Парсинг завершен")
        logger.info("=" * 60)
        logger.info("Всего строк: %s", self.stats.total_rows)
        logger.info("Документов: %s", self.stats.documents)
        logger.info("Пропущено: %s", self.stats.skipped_rows)
        logger.info("Ошибок: %s", self.stats.errors)
        logger.info("=" * 60)


    def parse_folder(self, folder: str) -> List[DocumentRecord]:
        """
        Парсит все Excel-файлы в папке.

        Args:
            folder: Путь к папке с файлами

        Returns:
            Список всех документов из всех файлов
        """
        result: List[DocumentRecord] = []
        folder = Path(folder)
        files = sorted(folder.glob("*.xlsx"))

        logger.info("Найдено файлов: %s", len(files))

        for file in files:
            logger.info("-" * 60)
            logger.info(file.name)
            documents = self.parse(str(file))
            result.extend(documents)
            self.print_statistics()

        return result